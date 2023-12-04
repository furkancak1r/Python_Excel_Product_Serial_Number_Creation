import os
import openpyxl

# Specify the file path
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\seri_no_create.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    try:
        # Load the workbook and the worksheets
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["2. Aşama"]
        result_sheet = workbook["3. Aşama"]

        # Get the used range of the worksheet
        rows = worksheet.max_row
        cols = worksheet.max_column

        # Access data and perform operations
        result_row = 1 # Initialize the result row counter
        for row in range(2, rows + 1):  # Assuming data starts from the second row
            cari_hesap_no = worksheet.cell(row=row, column=1).value
            cari_hesap_unvani = worksheet.cell(row=row, column=2).value
            sevkiyat_kodu= worksheet.cell(row=row, column=3).value
            sevkiyat_aciklamasi = worksheet.cell(row=row, column=4).value
            sevkiyat_adresi = worksheet.cell(row=row, column=5).value
            kodu= worksheet.cell(row=row, column=6).value
            aciklamasi= worksheet.cell(row=row, column=7).value
            miktar= worksheet.cell(row=row, column=8).value
            birim= worksheet.cell(row=row, column=9).value
            fatura_numarasi= worksheet.cell(row=row, column=10).value
            tarihi= worksheet.cell(row=row, column=11).value
            siparis_no= worksheet.cell(row=row, column=12).value
            gruplar= worksheet.cell(row=row, column=13).value
            

            # Perform your operations, for example, creating Dolap Seri No
            for i in range(int(miktar)):
                dolap_seri_no = f"{siparis_no}_{kodu}_{str(i+1).zfill(6)}" # Use i+1 instead of i and format it as a 6-character string with leading zeros
                result_sheet.cell(row=result_row, column=1).value = cari_hesap_no # Write the result to the result sheet
                result_sheet.cell(row=result_row, column=2).value = cari_hesap_unvani
                result_sheet.cell(row=result_row, column=3).value = sevkiyat_kodu
                result_sheet.cell(row=result_row, column=4).value = sevkiyat_aciklamasi
                result_sheet.cell(row=result_row, column=5).value = sevkiyat_adresi
                result_sheet.cell(row=result_row, column=6).value = kodu
                result_sheet.cell(row=result_row, column=7).value = aciklamasi
                result_sheet.cell(row=result_row, column=8).value = 1
                result_sheet.cell(row=result_row, column=9).value = birim
                result_sheet.cell(row=result_row, column=10).value = fatura_numarasi
                result_sheet.cell(row=result_row, column=11).value = tarihi
                result_sheet.cell(row=result_row, column=12).value = siparis_no
                result_sheet.cell(row=result_row, column=13).value = gruplar
                result_sheet.cell(row=result_row, column=14).value = dolap_seri_no
                
                result_row += 1 # Increment the result row counter

        # Save and close the workbook
        workbook.save(file_path)
        workbook.close()

    except Exception as e:
        print(f"An error occurred: {e}")
