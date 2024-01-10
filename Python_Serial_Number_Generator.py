import os
import openpyxl
# Seri no oluşturma
# Specify the file path
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\yazilacak.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    try:
        # Load the workbook and the worksheets
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["2"]
        result_sheet = workbook["3"]

        # Get the used range of the worksheet
        rows = worksheet.max_row
        cols = worksheet.max_column

        # Add headers to the result sheet
        result_sheet.cell(row=1, column=1).value = "Garanti Durumu"
        result_sheet.cell(row=1, column=2).value = "Fatura Tarihi / Garanti Başlangıç Tarihi"
        result_sheet.cell(row=1, column=3).value = "Fatura Numarası"
        result_sheet.cell(row=1, column=4).value = "Garanti Bitiş Tarihi"
        result_sheet.cell(row=1, column=5).value = "Sipariş No"
        result_sheet.cell(row=1, column=6).value = "Cari Hesap Kodu"
        result_sheet.cell(row=1, column=7).value = "Cari Hesap Unvanı"
        result_sheet.cell(row=1, column=8).value = "Cari Şube Mağaza / Sevk Adresleri"
        result_sheet.cell(row=1, column=9).value = "SEVKİYAT ADRESİ"
        result_sheet.cell(row=1, column=10).value = "Stok Kodu"
        result_sheet.cell(row=1, column=11).value = "Açıklaması"
        result_sheet.cell(row=1, column=12).value = "Satır Açıklaması"
        result_sheet.cell(row=1, column=13).value = "Miktarı"
        result_sheet.cell(row=1, column=14).value = "Birim"
        result_sheet.cell(row=1, column=15).value = "SERİ NUMARASI"


        # Access data and perform operations
        result_row = 2 # Initialize the result row counter
        for row in range(2, rows + 1):  # Assuming data starts from the second row
            Garanti_Durumu = worksheet.cell(row=row, column=1).value
            fatura_tarihi = worksheet.cell(row=row, column=2).value
            fatura_numarasi = worksheet.cell(row=row, column=3).value
            garanti_bitis_tarihi = worksheet.cell(row=row, column=4).value
            Sipariş_No = worksheet.cell(row=row, column=5).value
            cari_hesap_no = worksheet.cell(row=row, column=6).value
            cari_hesap_unvani = worksheet.cell(row=row, column=7).value
            cari_sube_magaza_sevk_adresleri = worksheet.cell(row=row, column=8).value
            sevkiyat_adresi = worksheet.cell(row=row, column=9).value
            Stok_Kodu = worksheet.cell(row=row, column=10).value
            aciklamasi = worksheet.cell(row=row, column=11).value
            satir_aciklamasi = worksheet.cell(row=row, column=12).value
            Miktarı = worksheet.cell(row=row, column=13).value
            Birim = worksheet.cell(row=row, column=14).value
            seri_numarasi = worksheet.cell(row=row, column=15).value
            if Miktarı is not None and Birim == "ADET":  # Check if Miktarı is not None
                # Perform your operations, for example, creating Dolap Seri No
                for i in range(int(Miktarı)):
                    
                    #dolap_seri_no=f"{Sipariş_No}_{Stok_Kodu}_{000000}" 
                    dolap_seri_no = f"{Sipariş_No}_{Stok_Kodu}_{str(i+1).zfill(6)}" # Use i+1 instead of i and format it as a 6-character string with leading zeros
                    result_sheet.cell(row=result_row, column=1).value = Garanti_Durumu # Write the result to the result sheet
                    result_sheet.cell(row=result_row, column=2).value = fatura_tarihi
                    result_sheet.cell(row=result_row, column=3).value = fatura_numarasi
                    result_sheet.cell(row=result_row, column=4).value = garanti_bitis_tarihi
                    result_sheet.cell(row=result_row, column=5).value = Sipariş_No
                    result_sheet.cell(row=result_row, column=6).value = cari_hesap_no
                    result_sheet.cell(row=result_row, column=7).value = cari_hesap_unvani
                    result_sheet.cell(row=result_row, column=8).value = cari_sube_magaza_sevk_adresleri
                    result_sheet.cell(row=result_row, column=9).value = sevkiyat_adresi
                    result_sheet.cell(row=result_row, column=10).value = Stok_Kodu
                    result_sheet.cell(row=result_row, column=11).value = aciklamasi
                    result_sheet.cell(row=result_row, column=12).value = satir_aciklamasi
                    result_sheet.cell(row=result_row, column=13).value = 1
                    result_sheet.cell(row=result_row, column=14).value = Birim
                    result_sheet.cell(row=result_row, column=15).value = seri_numarasi
                    result_sheet.cell(row=result_row, column=16).value = dolap_seri_no
                    
                    
                    result_row += 1 # Increment the result row counter
            elif Miktarı is not None: 
                    Miktarı_tam = int(Miktarı)

                    # Dolap seri numarasını formatlama
                    dolap_seri_no = f"{Sipariş_No}_{Stok_Kodu}_{Miktarı_tam:06d}"
                    #dolap_seri_no=f"{Sipariş_No}_{Stok_Kodu}_{'000000'}" 
                    result_sheet.cell(row=result_row, column=1).value = Garanti_Durumu # Write the result to the result sheet
                    result_sheet.cell(row=result_row, column=2).value = fatura_tarihi
                    result_sheet.cell(row=result_row, column=3).value = fatura_numarasi
                    result_sheet.cell(row=result_row, column=4).value = garanti_bitis_tarihi
                    result_sheet.cell(row=result_row, column=5).value = Sipariş_No
                    result_sheet.cell(row=result_row, column=6).value = cari_hesap_no
                    result_sheet.cell(row=result_row, column=7).value = cari_hesap_unvani
                    result_sheet.cell(row=result_row, column=8).value = cari_sube_magaza_sevk_adresleri
                    result_sheet.cell(row=result_row, column=9).value = sevkiyat_adresi
                    result_sheet.cell(row=result_row, column=10).value = Stok_Kodu
                    result_sheet.cell(row=result_row, column=11).value = aciklamasi
                    result_sheet.cell(row=result_row, column=12).value = satir_aciklamasi
                    result_sheet.cell(row=result_row, column=13).value = Miktarı
                    result_sheet.cell(row=result_row, column=14).value = Birim
                    result_sheet.cell(row=result_row, column=15).value = seri_numarasi
                    result_sheet.cell(row=result_row, column=16).value = dolap_seri_no
                    
                    result_row += 1
            else: pass    
                
                
        # Save and close the workbook
        workbook.save(file_path)
        workbook.close()

    except Exception as e:
        print(f"An error occurred: {e}")
