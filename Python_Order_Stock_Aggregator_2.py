import openpyxl

# Load the workbook and worksheets
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\26 ARALIK 2023 VE OCAK 2024 SATIŞ FATURALARI  ÜZERİNDEN - GARANTİLİ ÜRÜNLER ÇALIŞMASI.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook["1"]
result_sheet = workbook["2"]

# Clear existing data in result_sheet
result_sheet.delete_rows(2, result_sheet.max_row + 1)

# Process and combine rows
combined_data = {}
for row in range(2, worksheet.max_row + 1):
    order_no = worksheet.cell(row=row, column=6).value  # Sipariş No
    stock_code = worksheet.cell(row=row, column=7).value  # Stok Kodu
    quantity = worksheet.cell(row=row, column=9).value  # Miktar
    # Convert None to 0 and handle commas in quantity
    quantity = 0 if quantity is None else quantity

    key = (order_no, stock_code)
    if key not in combined_data:
        combined_data[key] = [worksheet.cell(row=row, column=i).value for i in range(1, worksheet.max_column + 1)]
        # Ensure the list has enough items
        while len(combined_data[key]) <= 12:
            combined_data[key].append(None)
            print("quantity: ", quantity, "combined_data[key][11]: ", combined_data[key][11])
        combined_data[key][12] = quantity  # Set quantity
        combined_data[key][11] = None  # Clear Açıklama
    else:
        combined_data[key][12] += quantity  # Add to quantity
        # Print order_no and stock_code when quantities are combined

# Write combined data to result sheet
result_row = 2
for key, values in combined_data.items():
    for col, value in enumerate(values, start=1):
        result_sheet.cell(row=result_row, column=col).value = value
    result_row += 1

# Save the workbook
workbook.save(file_path)
