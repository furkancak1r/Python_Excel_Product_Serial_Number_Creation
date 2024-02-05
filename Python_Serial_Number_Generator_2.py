import os
import openpyxl
# Seri no oluşturma
# Specify the file path
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\26 ARALIK 2023 VE OCAK 2024 SATIŞ FATURALARI  ÜZERİNDEN - GARANTİLİ ÜRÜNLER ÇALIŞMASI.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    try:
        # Load the workbook and the worksheets
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["2"]
        result_sheet = workbook["3"]

        # Get the used range of the worksheet
        rows = worksheet.max_row
        cols = worksheet.max_column

        # Add headers to the result sheet
        result_sheet.cell(row=1, column=1).value = "GARANTİ BAŞLANGIÇ TARİHİ  tarihi"
        result_sheet.cell(row=1, column=2).value = "GARANTİ BİTİŞ TARİHİ"
        result_sheet.cell(row=1, column=3).value = "MUHATAP"
        result_sheet.cell(row=1, column=4).value = "MAĞAZA ADI"
        result_sheet.cell(row=1, column=5).value = "SERİ NO"
        result_sheet.cell(row=1, column=6).value = "SATIŞ SİP. NO"
        result_sheet.cell(row=1, column=7).value = "Kalem Kodu"
        result_sheet.cell(row=1, column=8).value = "Kalem Tanımı"
        result_sheet.cell(row=1, column=9).value = "Miktar"
        result_sheet.cell(row=1, column=10).value = "Birim"
        result_sheet.cell(row=1, column=11).value = "Sevkiyat Adresi"



        # Access data and perform operations
        result_row = 2 # Initialize the result row counter
        for row in range(2, rows + 1):  # Assuming data starts from the second row
            garanti_baslangic = worksheet.cell(row=row, column=1).value
            garanti_bitis = worksheet.cell(row=row, column=2).value
            muhatap = worksheet.cell(row=row, column=3).value
            magaza_adi = worksheet.cell(row=row, column=4).value
            seri_no = worksheet.cell(row=row, column=5).value
            siparis_no = worksheet.cell(row=row, column=6).value
            kalem_kodu = worksheet.cell(row=row, column=7).value
            kalem_tanimi = worksheet.cell(row=row, column=8).value
            miktar = worksheet.cell(row=row, column=9).value
            Birim = worksheet.cell(row=row, column=10).value
            sevkiyat_adresi = worksheet.cell(row=row, column=11).value
            
          
            if miktar is not None and Birim == "ADET":  # Check if Miktarı is not None
                # Perform your operations, for example, creating Dolap Seri No
                for i in range(int(miktar)):
                    
                    #dolap_seri_no=f"{Sipariş_No}_{Stok_Kodu}_{000000}" 
                    dolap_seri_no = f"{siparis_no}_{kalem_kodu}_{str(i+1).zfill(6)}" # Use i+1 instead of i and format it as a 6-character string with leading zeros
                    result_sheet.cell(row=result_row, column=1).value = garanti_baslangic # Write the result to the result sheet
                    result_sheet.cell(row=result_row, column=2).value = garanti_bitis
                    result_sheet.cell(row=result_row, column=3).value = muhatap
                    result_sheet.cell(row=result_row, column=4).value = magaza_adi
                    result_sheet.cell(row=result_row, column=5).value = seri_no
                    result_sheet.cell(row=result_row, column=6).value = siparis_no
                    result_sheet.cell(row=result_row, column=7).value = kalem_kodu
                    result_sheet.cell(row=result_row, column=8).value = kalem_tanimi
                    result_sheet.cell(row=result_row, column=9).value = 1
                    result_sheet.cell(row=result_row, column=10).value = Birim
                    result_sheet.cell(row=result_row, column=11).value = sevkiyat_adresi
                    result_sheet.cell(row=result_row, column=12).value = dolap_seri_no
                    
                    
                    result_row += 1 # Increment the result row counter
            elif miktar is not None: 
                    Miktarı_tam = int(miktar)

                    # Dolap seri numarasını formatlama
                    dolap_seri_no = f"{siparis_no}_{kalem_kodu}_{Miktarı_tam:06d}"
                    #dolap_seri_no=f"{Sipariş_No}_{Stok_Kodu}_{'000000'}" 
                    result_sheet.cell(row=result_row, column=1).value = garanti_baslangic # Write the result to the result sheet
                    result_sheet.cell(row=result_row, column=2).value = garanti_bitis
                    result_sheet.cell(row=result_row, column=3).value = muhatap
                    result_sheet.cell(row=result_row, column=4).value = magaza_adi
                    result_sheet.cell(row=result_row, column=5).value = seri_no
                    result_sheet.cell(row=result_row, column=6).value = siparis_no
                    result_sheet.cell(row=result_row, column=7).value = kalem_kodu
                    result_sheet.cell(row=result_row, column=8).value = kalem_tanimi
                    result_sheet.cell(row=result_row, column=9).value = miktar
                    result_sheet.cell(row=result_row, column=10).value = Birim
                    result_sheet.cell(row=result_row, column=11).value = sevkiyat_adresi
                    result_sheet.cell(row=result_row, column=12).value = dolap_seri_no
                    result_row += 1
            else: pass    
                
                
        # Save and close the workbook
        workbook.save(file_path)
        workbook.close()

    except Exception as e:
        print(f"An error occurred: {e}")
