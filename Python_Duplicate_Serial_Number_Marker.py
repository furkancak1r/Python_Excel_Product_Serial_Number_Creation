import os
import openpyxl
# Tekrar edenleri yazıyor
# Specify the file path
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\yazilacak.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    try:
        # Load the workbook and the worksheet
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["3"]

        # Get the used range of the worksheet
        rows = worksheet.max_row

        # Collect all seri numarasi values
        seri_numarasi_values = [worksheet.cell(row=row, column=15).value for row in range(2, rows + 1)]

        # Identify duplicates
        duplicates = set([x for x in seri_numarasi_values if seri_numarasi_values.count(x) > 1])

        # Mark duplicates in the worksheet
        for row in range(2, rows + 1):
            cell_value = worksheet.cell(row=row, column=15).value
            if cell_value in duplicates:
                worksheet.cell(row=row, column=16).value = "Tekrar ediyor"

        # Save and close the workbook
        workbook.save(file_path)
        workbook.close()

    except Exception as e:
        print(f"An error occurred: {e}")
